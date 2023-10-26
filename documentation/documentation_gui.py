# GUI 관련 패키지
import tkinter.ttk as ttk
from tkinter import *
from tkinter import filedialog

#hanaDB 연결 및 엑셀 패키지
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
import pandas as pd
from datetime import datetime
from hdbcli import dbapi

root = Tk()
root.title("테이블정의서 일괄 생성")

# 저장 경로 (폴더)
def browse_dest_path():
    folder_selected = filedialog.askdirectory()
    if folder_selected is None: # 사용자가 취소를 누를 때
        return
    dest_path.delete(0, END)
    dest_path.insert(0, folder_selected)

# 테이블 정의서 생성
def documentation(df, username, curdate):
    # 엑셀문서 열기
    workbook = load_workbook(filename='_documentation_template.xlsx')
    # 스타일 생성
    border_side = Side(border_style='thin')
    square_border = Border(top=border_side,
                        right=border_side,
                        bottom=border_side,
                        left=border_side)
    align_center = Alignment(horizontal="center")
    # 첫장 SHEET 선택
    sheet = workbook['첫장']
    # 첫장 변경이력 생성
    sheet.cell(row=7, column=3).value = curdate
    sheet.cell(row=7, column=9).value = username
    # 스타일 적용
    for row in range(6, 24):
        for col in range(1, 10):
            sheet.cell(row=row, column=col).border = square_border
    # 테이블정의서 SHEET 선택
    sheet = workbook['테이블정의서']
    # 스타일 적용
    for row in range(2, 6):
        for col in range(2, 8):
            sheet.cell(row=row, column=col).border = square_border
    # Table Header 정보 지정
    sheet.cell(row=3, column=4).value = df.iloc[0, 1]+'.'+df.iloc[0, 2]             # 테이블ID
    sheet.cell(row=3, column=6).value = df.iloc[0, 3]                               # 테이블명
    sheet.cell(row=4, column=4).value = df.iloc[0, 3]                               # 테이블개요
    sheet.cell(row=4, column=6).value = 'v0.1'                                      # 문서버전
    sheet.cell(row=5, column=4).value = username                                    # 작성자
    sheet.cell(row=5, column=6).value = curdate                                     # 작성일
    # Table Item 정보 지정
    for num, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        sheet.cell(row=8+num, column=2).value = row[4]                              # Field
        sheet.cell(row=8+num, column=3).value = row[5]                              # Key
        sheet.cell(row=8+num, column=4).value = row[6]                              # Desc
        sheet.cell(row=8+num, column=5).value = row[7]                              # Type
        sheet.cell(row=8+num, column=6).value = row[8]                              # Size
        # Table Item 스타일 지정
        sheet.cell(row=8+num, column=6).alignment = align_center
        for col in range(2, 8):
            sheet.cell(row=8+num, column=col).border = square_border
    # 엑셀 저장
    workbook.save(filename='D11_DI_테이블정의서_'+df.iloc[0, 2]+'_v0.1.xlsx')

def start():
    # 저장경로 지정 여부 점검
    if dest_path.get() is "":
        list_file.insert(0, "저장경로를 지정하세요!")
        return

    # 기본 작업 디렉토리 변경
    os.chdir(dest_path.get())
    # HANA DB 연결 후 테이블 목록과 레이아웃을 조회하여 테이블정의서 산출물 생성
    # HANA DB 연결 
    cnxn = dbapi.connect("10.1.61.41", 30041, "ztpm_dw", "Ztpm_dw1234#$")
    # 테이블 목록 조회 쿼리
    tablelist = """
    SELECT DISTINCT
        OBJ_TYPE
        , SCHEMA_NAME
        , TABLE_NAME
        , TABLE_DESC
    FROM COLS
    WHERE 1=1
        AND OBJ_TYPE = 'TABLE'
        AND SCHEMA_NAME = 'ZTPM_DW'
    """
    # 테이블 목록을 데이터 프레임에 저장
    dflist = pd.read_sql(tablelist, cnxn, index_col=None)
    # 테이블 목록별로 루프 실행
    for num, row in enumerate(dataframe_to_rows(dflist, index=False, header=False)):
        # 테이블별 상세 정의 쿼리
        tabledetail = """
            SELECT 
                OBJ_TYPE
                , SCHEMA_NAME
                , TABLE_NAME
                , TABLE_DESC
                , COLUMN_NAME
                , PK
                , COLUMN_DESC
                , DATA_TYPE
                , LEN_SCALE
            FROM COLS
            WHERE 1=1
                AND OBJ_TYPE = 'TABLE'
                AND TABLE_NAME = '%s'
            ORDER BY NO
        """ % row[2]
        # 테이블 레이아웃 조회
        df = pd.read_sql(tabledetail, cnxn, index_col=None)
        
        # progressbar 업데이트
        progress = (num+1) / len(dflist) * 100
        p_var.set(progress)
        progress_bar.update()
        list_file.insert(0, row[2])
        # 문서생성 시작
        documentation(df, writer_entry.get(), date_entry.get())
    list_file.insert(0, "테이블정의서 생성 완료!")

def popup():
    toplevel=Toplevel(root)
    toplevel.title("About")
    toplevel.geometry("200x100")
    label=Label(toplevel, text=history, justify="left")
    label.pack()

# 변경이력
history = """
작성자 : 윤치선
작성일 : 2020.06.29

version : 1.0
"""

# 메뉴생성
menubar = Menu(root)
menu_1 = Menu(menubar, tearoff=0)
menu_1.add_command(label="About", command=popup)
menubar.add_cascade(label="About", menu=menu_1)
root.config(menu=menubar)

# 작성자 프레임
write_frame = Frame(root)
write_frame.pack(fill="x", padx=5, pady=5) # 간격 띄우기

label1 = Label(write_frame, text="작성자")
label1.pack(side="left")

writer_entry = Entry(write_frame)
writer_entry.pack(side="right")
writer_entry.insert(0, "윤치선")

# 작성일 프레임
date_frame = Frame(root)
date_frame.pack(fill="x", padx=5, pady=5) # 간격 띄우기

label2 = Label(date_frame, text="작성일")
label2.pack(side="left")

date_entry = Entry(date_frame)
date_entry.pack(side="right")

# 현재시간
now = datetime.now()
curdate = '%s-%s-%s' % (now.year, now.month, now.day)
date_entry.insert(0, curdate)

# 저장 경로 프레임
path_frame = LabelFrame(root, text="저장경로")
path_frame.pack(fill="x", padx=5, pady=5, ipady=5)

dest_path = Entry(path_frame)
dest_path.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4) # 높이 변경

btn_dest_path = Button(path_frame, text="찾아보기", width=10, command=browse_dest_path)
btn_dest_path.pack(side="right", padx=5, pady=5)

# 진행 상황 Progress Bar
frame_progress = LabelFrame(root, text="진행상황")
frame_progress.pack(fill="x", padx=5, pady=5, ipady=5)

p_var = DoubleVar()
progress_bar = ttk.Progressbar(frame_progress, maximum=100, variable=p_var)
progress_bar.pack(fill="x", padx=5, pady=5)

# 실행 프레임
frame_run = Frame(root)
frame_run.pack(fill="x", padx=5, pady=5)

btn_close = Button(frame_run, padx=5, pady=5, text="닫기", width=12, command=root.quit)
btn_close.pack(side="right", padx=5, pady=5)

btn_start = Button(frame_run, padx=5, pady=5, text="시작", width=12, command=start)
btn_start.pack(side="right", padx=5, pady=5)

# 리스트 프레임
list_frame = Frame(root)
list_frame.pack(fill="both", padx=5, pady=5)

scrollbar = Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")

list_file = Listbox(list_frame, selectmode="extended", height=15, yscrollcommand=scrollbar.set, \
    background="black", foreground="white")
list_file.pack(side="left", fill="both", expand=True)
scrollbar.config(command=list_file.yview)

root.resizable(False, False)

root.mainloop()
