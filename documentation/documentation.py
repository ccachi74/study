# HANA DB 연결 후 테이블 목록과 레이아웃을 조회하여 테이블정의서 산출물 생성

import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
import pandas as pd
from datetime import datetime
from hdbcli import dbapi

# HANA DB 연결 
cnxn = dbapi.connect("10.1.61.41", 30041, "ztpm_dw", "Ztpm_dw1234#$")

# 기본 작업 디렉토리 변경
os.chdir("documentation")

# 테이블 목록 조회 쿼리
tablelist = """
SELECT DISTINCT
    OBJ_TYPE
    , SCHEMA_NAME
    , TABLE_NAME
    , TABLE_DESC
FROM COLS
WHERE 1=1
    AND OBJ_TYPE = 'TABLE'
    AND SCHEMA_NAME = 'ZTPM_DW'
"""

# 테이블 목록을 데이터 프레임에 저장
dflist = pd.read_sql(tablelist, cnxn, index_col=None)

# 현재시간
now = datetime.now()
curdate = '%s-%s-%s' % (now.year, now.month, now.day)

# 작성자명
username = input('작성자명 : ')

# 테이블 정의서 생성
def documentation(df, username, curdate):
    # 엑셀문서 열기
    workbook = load_workbook(filename='D11_DI_테이블정의서_ZT_SAMPLE_v0.1.xlsx')

    # 스타일 생성
    border_side = Side(border_style='thin')
    square_border = Border(top=border_side,
                           right=border_side,
                           bottom=border_side,
                           left=border_side)

    align_center = Alignment(horizontal="center")

    # 첫장 SHEET 선택
    sheet = workbook['첫장']

    # 첫장 변경이력 생성
    sheet.cell(row=7, column=3).value = curdate
    sheet.cell(row=7, column=9).value = username

    # 스타일 적용
    for row in range(6, 24):
        for col in range(1, 10):
            sheet.cell(row=row, column=col).border = square_border

    # 테이블정의서 SHEET 선택
    sheet = workbook['테이블정의서']

    # 스타일 적용
    for row in range(2, 6):
        for col in range(2, 8):
            sheet.cell(row=row, column=col).border = square_border

    # Table Header 정보 지정
    sheet.cell(row=3, column=4).value = df.iloc[0, 1]+'.'+df.iloc[0, 2]             # 테이블ID
    sheet.cell(row=3, column=6).value = df.iloc[0, 3]                               # 테이블명
    sheet.cell(row=4, column=4).value = df.iloc[0, 3]                               # 테이블개요
    sheet.cell(row=4, column=6).value = 'v0.1'                                      # 문서버전
    sheet.cell(row=5, column=4).value = username                                    # 작성자
    sheet.cell(row=5, column=6).value = curdate                                     # 작성일

    # Table Item 정보 지정
    for num, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
        sheet.cell(row=8+num, column=2).value = row[4]                              # Field
        sheet.cell(row=8+num, column=3).value = row[5]                              # Key
        sheet.cell(row=8+num, column=4).value = row[6]                              # Desc
        sheet.cell(row=8+num, column=5).value = row[7]                              # Type
        sheet.cell(row=8+num, column=6).value = row[8]                              # Size

        # Table Item 스타일 지정
        sheet.cell(row=8+num, column=6).alignment = align_center
        for col in range(2, 8):
            sheet.cell(row=8+num, column=col).border = square_border

    # 엑셀 저장
    workbook.save(filename='D11_DI_테이블정의서_'+df.iloc[0, 2]+'_v0.1.xlsx')

# 테이블 목록별로 루프 실행
for num, row in enumerate(dataframe_to_rows(dflist, index=False, header=False)):
    # 테이블별 상세 정의 쿼리
    tabledetail = """
        SELECT 
            OBJ_TYPE
            , SCHEMA_NAME
            , TABLE_NAME
            , TABLE_DESC
            , COLUMN_NAME
            , PK
            , COLUMN_DESC
            , DATA_TYPE
            , LEN_SCALE
        FROM COLS
        WHERE 1=1
            AND OBJ_TYPE = 'TABLE'
            AND TABLE_NAME = '%s'
        ORDER BY NO
    """ % row[2]

    # 테이블 레이아웃 조회
    df = pd.read_sql(tabledetail, cnxn, index_col=None)
    print('테이블 %s 를 생성합니다.' % row[2])
    documentation(df, username, curdate)

